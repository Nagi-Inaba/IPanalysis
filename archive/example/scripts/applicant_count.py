import openpyxl
from collections import Counter
import sys 


# ワークブックをロード
fname = sys.argv[1]

# 出願年の範囲を入力
start_year = int(sys.argv[2])
end_year = int(sys.argv[3])

# Excelファイルを開く
wb = openpyxl.load_workbook(fname)
sheet = wb['データ']

# 筆頭出願人と出願年の列を探す
applicant_column = None
year_column = None

for col in sheet.iter_cols(1, sheet.max_column):
    if col[0].value == '筆頭出願人':
        applicant_column = col[0].column
    elif col[0].value == '出願年':
        year_column = col[0].column
    
    if applicant_column and year_column:
        break

if applicant_column is not None and year_column is not None:
    # 出願年と筆頭出願人のデータをカウント
    applicants = []
    for row in sheet.iter_rows(min_row=2, min_col=min(applicant_column, year_column), max_col=max(applicant_column, year_column)):
        applicant_value = row[applicant_column - 1].value
        year_value = row[year_column - 1].value
        
        # 出願年が指定範囲内か確認
        if isinstance(year_value, int) and start_year <= year_value <= end_year:
            if applicant_value is not None:
                applicants.append(applicant_value)

    # 出願件数の集計
    applicant_counts = Counter(applicants)

    # 集計結果を出願件数が多い順に並べ替え
    sorted_applicants = sorted(applicant_counts.items(), key=lambda x: x[1], reverse=True)

    # 新しいシートを作成し、結果を記述
    result_sheet = wb.create_sheet(title="筆頭出願人件数集計")
    result_sheet.cell(row=1, column=1, value='筆頭出願人')
    result_sheet.cell(row=1, column=2, value='出願件数')

    row_index = 2
    for applicant, count in sorted_applicants:
        result_sheet.cell(row=row_index, column=1, value=applicant)
        result_sheet.cell(row=row_index, column=2, value=count)
        row_index += 1

    # 結果を保存
    wb.save(fname)
    print("集計結果が新しいシートに書き込まれました。")

else:
    print('「筆頭出願人」または「出願年」列が見つかりませんでした。')

