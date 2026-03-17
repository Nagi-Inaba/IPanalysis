import openpyxl
from collections import Counter
import sys 


# ワークブックをロード
fname = sys.argv[1]
sheet_name = 'データ'  # シート名は文字列で指定

def aggregate_data_and_copy_to_new_sheet(fname, sheet_name, year_column_header):
    # ファイルをロード
    wb = openpyxl.load_workbook(fname)

    # シートを選択
    sheet = wb[sheet_name]

    # 出願年の列を探す
    year_col = None
    for col in sheet.iter_cols(1, sheet.max_column):
        if col[0].value == year_column_header:
            year_col = col[0].column
            break

    if year_col is not None:
        # 出願年データのカウント
        years = []
        for row in sheet.iter_rows(min_row=2, min_col=year_col, max_col=year_col):
            year_value = row[0].value
            if year_value is not None:
                years.append(year_value)

        year_counts = Counter(years)

        # 新しいシートを作成し、名前を設定
        chart_sheet = wb.create_sheet(title="出願件数")

        # ヘッダー行を設定
        chart_sheet.cell(row=1, column=1, value='出願年')
        chart_sheet.cell(row=1, column=2, value='出願件数')

        # 集計結果を新しいシートに追加
        row_index = 2
        for year, count in sorted(year_counts.items()):
            chart_sheet.cell(row=row_index, column=1, value=year)
            chart_sheet.cell(row=row_index, column=2, value=count)
            row_index += 1

    # ファイルを保存
    wb.save(fname)
    wb.close()

# 実行
aggregate_data_and_copy_to_new_sheet(fname, 'データ', '出願年')