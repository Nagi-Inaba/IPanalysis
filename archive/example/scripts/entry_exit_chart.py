import openpyxl
from collections import defaultdict
import sys 


# ワークブックをロード
fname = sys.argv[1]

# Excelファイルを開く
wb = openpyxl.load_workbook(fname)
sheet = wb['データ']

# 更新出願人権利者と出願年の列を探す
applicant_column = None
year_column = None

# 列番号の取得（注意: 列名が正しいか確認）
for col in sheet.iter_cols(1, sheet.max_column):
    if col[0].value == '更新出願人・権利者氏名':
        applicant_column = col[0].column
    elif col[0].value == '出願年':
        year_column = col[0].column

    if applicant_column and year_column:
        break

# エラーを回避するためのチェック
if applicant_column is None:
    raise ValueError('「更新出願人・権利者氏名」列が見つかりません。')
if year_column is None:
    raise ValueError('「出願年」列が見つかりません。')

# 出願年と更新出願人権利者のデータを集計
data = defaultdict(lambda: {"first_year": float('inf'), "last_year": -float('inf'), "count": 0})

# データを集計する（範囲指定を確認）
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=min(applicant_column, year_column), max_col=max(applicant_column, year_column)):
    applicant_value = row[applicant_column - min(applicant_column, year_column)].value
    year_value = row[year_column - min(applicant_column, year_column)].value

    if isinstance(year_value, int) and applicant_value is not None:
        # コンマで区切られた出願人を分割し、重複を排除
        applicants = set([applicant.strip() for applicant in applicant_value.split(',')]) # type: ignore
        for applicant in applicants:
            # 最初の出願年、直近の出願年、出願件数を集計
            data[applicant]["first_year"] = min(data[applicant]["first_year"], year_value)
            data[applicant]["last_year"] = max(data[applicant]["last_year"], year_value)
            data[applicant]["count"] += 1

# 新しいシートを作成し、結果を記述
result_sheet = wb.create_sheet(title="参入撤退チャート")

# ヘッダーの設定
result_sheet.cell(row=1, column=1, value='更新出願人・権利者氏名')
result_sheet.cell(row=1, column=2, value='最初の出願年')
result_sheet.cell(row=1, column=3, value='直近出願年')
result_sheet.cell(row=1, column=4, value='総出願件数')

# 各出願人のデータの書き込み
row_index = 2
for applicant, info in data.items():
    result_sheet.cell(row=row_index, column=1, value=applicant)
    result_sheet.cell(row=row_index, column=2, value=info["first_year"])
    result_sheet.cell(row=row_index, column=3, value=info["last_year"])
    result_sheet.cell(row=row_index, column=4, value=info["count"])
    row_index += 1

# 結果を保存
wb.save(fname)
print("集計結果が新しいシートに書き込まれました。")
