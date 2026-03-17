import openpyxl
from datetime import datetime, timedelta
import unicodedata
import pandas as pd
import warnings
import sys  # 追加

# 警告を非表示にする
warnings.filterwarnings("ignore", category=UserWarning, module='openpyxl')

# ★ここを変更：input() → 引数
fname = sys.argv[1]
wb = openpyxl.load_workbook(fname)
# ワークシート名を選択
sheet = wb['データ']  


# ヘッダーとして探す値
header_value = '更新出願人・権利者氏名'
target_col = None

# 最上部のセルをチェックして、目的の列を見つける
for col in sheet.iter_cols(min_row=1, max_row=1):
    if col[0].value == header_value:
        target_col = col[0].column
        break


# 特定の文字列を指定
target_chars = ['株式会社','合同会社', 'カブシキガイシャ', 'カブシキガイシヤ','有限会社', '有限公司', '(株)', 'コーポレーション', "特定非営利活動法人","（株）","公立大学法人","財団法人","財團法人",'コムパニー',
                 'インコーポレイテッド', 'インコーポレーテッド', 'エルエルシー', '独立行政法人','ホールディングス','ホールディング','一般社団法人','一般財団法人',"合資会社","合名会社","リミテッド","エルティーディ","アーゲー","ホウルディング",
                 '社会福祉法人', '国立大学法人', '国立研究開発法人', '国立研究開発機構','公益財団法人','カンパニイ','インコ－ポレイテッド','ゲーエムベーハー','コーポレーシヨン',"社団法人",'コーポレイション','有限責任公司',
                 '学校法人','公益社団法人' ,'オーガナイゼーション', 'コーポレイシヨン','コーポレイション' ,'コ－ポレ－シヨン', 'カンパニー', 'システムズ', 'インク', 'インスティテュート',' ','　','・','.','．','、','，','▼','▲',
                 ]


for cell in sheet.iter_rows(min_col=target_col, max_col=target_col, min_row=2):
        cell = cell[0]

        if cell.value is not None:
            if isinstance(cell.value, str):
                # セルの値が文字列の場合、複数の文字列を順に置換する
                for target_char in target_chars:
                    cell.value = cell.value.replace(target_char, '')
            elif isinstance(cell.value, (list, tuple)):
                # セルの値がリストまたはタプルの場合、文字列に変換してから処理する
                if isinstance(cell.value, list):
                    new_value = ''.join(str(item) for item in cell.value if isinstance(item, str))
                elif isinstance(cell.value, tuple):
                    new_value = ''.join(str(item) for item in cell.value if isinstance(item, str)) # type: ignore
                # リストまたはタプルの要素を置換し、文字列に戻す
                new_value = ''.join(str(item).replace(target_char, '') for item in new_value)
                cell.value = new_value
            else:
                # セルの値が他の型の場合、そのままにする
                pass

# ヘッダー行を探索して「更新出願人・権利者氏名」列を自動で探す
column_name = "更新出願人・権利者氏名"
target_column = None

for col in range(1, sheet.max_column + 1):
    header = sheet.cell(row=1, column=col).value
    if header == column_name:
        target_column = col
        break

if target_column is None:
    print(f"列 '{column_name}' が見つかりませんでした。")
else:
    # 変換対象となる会社名と、変換後の名前の辞書
    replace_dict = {
        "－":"ー","—":"ー","ー":"ー",
        "三菱重工エンジニアリング": "三菱重工業","三菱パワー": "三菱重工業","三菱日立パワーシステムズ": "三菱重工業","三菱造船": "三菱重工業","三菱重工パワーインダストリー": "三菱重工業",
        "三菱重工エンジンアンドターボチャージャ": "三菱重工業","三菱重工マリンマシナリ": "三菱重工業","三菱エフビーアールシステムズ": "三菱重工業","三菱原子燃料": "三菱重工業","三菱重工サーマル":"三菱重工業",
        "三菱重工航空エンジン": "三菱重工業","三菱ロジスネクスト": "三菱重工業","三菱重工交通建設エンジニアリング": "三菱重工業","三菱重工パワー環境ソリューション":"三菱重工業","三菱重工機械システム": "三菱重工業",
        "三菱重工サーマルシステムズ": "三菱重工業","ＭＨＩＮＳエンジニアリング": "三菱重工業","ＭＨＩ原子力研究開発": "三菱重工業","ＭＨＩソリューションテクノロジーズ": "三菱重工業","ＭＨＩパワーコントロールシステムズ": "三菱重工業","ＭＨＩパワーエンジニアリング": "三菱重工業",
        "三菱日立パワー": "三菱重工業","テクノ電子": "三菱重工業","三菱ＦＢＲシステムズ": "三菱重工業","ＭＨＩエアロスペースシステムズ": "三菱重工業","中菱エンジニアリング": "三菱重工業","ＭＨＩ下関エンジニアリング": "三菱重工業",
        "ＭＨＩＬＮＧ": "三菱重工業","エムエイチアイさがみハイテック": "三菱重工業","ニチユマシナリー": "三菱重工業","ＭＨＩハセック": "三菱重工業","ＭＨＩマリテック": "三菱重工業","ＭＨＩエンジニアリング": "三菱重工業","三菱重工環境化学エンジニアリング": "三菱重工業",
        "三菱重工冷熱": "三菱重工業","三菱重工コンプレッサ": "三菱重工業","三菱マヒンドラ農機":"三菱重工業","三菱農機":"三菱重工業","バブコツク日立":"三菱重工業",
        "三菱ケミカルエンジニアリング":"三菱ケミカル","三菱ケミカルアグリドリーム":"三菱ケミカル","三菱レイヨン":"三菱ケミカル","三菱化学":"三菱ケミカル","三菱化成":"三菱ケミカル","三菱化成ビニル":"三菱ケミカル",
        "三菱鉱業セメント":"三菱マテリアル","三菱電線工業":"三菱マテリアル","三菱マテリアル電子化成":"三菱マテリアル","三菱樹脂":"三菱ケミカル","三菱ケミカルポリエステルフィルム":"三菱ケミカル",
        "三菱電機エンジニアリング":"三菱電機","三菱油化":"三菱ケミカル","三菱ケミカルビニル":"三菱ケミカル","三菱ケミカルポリテツク":"三菱ケミカル","三菱ケミカルポリテック":"三菱ケミカル",
        "パナソニックエコ":"パナソニック","松下電器": "パナソニック", "パナソニック産業": "パナソニック","パナソニック電工": "パナソニック","三洋電機": "パナソニック","パナソニックＥＶエナジー": "プライムアースＥＶエナジー",
        "松下電工": "パナソニック","松下電器産業": "パナソニック","松下電子": "パナソニック","パナソニックＩＰマネジメント": "パナソニック","パナソニックエナジー": "パナソニック",
        "新日本製鐵":"日本製鉄","新日本製鉄":"日本製鉄","新日鉄住金":"日本製鉄","住友金属工業":"日本製鉄","日新製鋼":"日本製鉄","日本製鉄マテリアルズ":"日本製鉄","新日鐵住金":"日本製鉄","日鉄ケミカル＆マテリアル":"日本製鉄","日鉄日本製鉄":"日本製鉄","日鉄ステンレス":"日本製鉄",
        "東芝エネルギーシステムズ": "東芝","東芝エネルギー": "東芝","東芝テクニカルサービスインターナショナル": "東芝","東芝セラミックス": "東芝","東芝アイテック": "東芝","東芝プラントシステム": "東芝",
        "東芝クライアントソリューション": "東芝","東芝システムテクノロジー": "東芝","東芝テック": "東芝","東芝プラント建設": "東芝","東芝電池": "東芝","東芝情報システム":"東芝","東芝デジタルソリューションズ":"東芝","東芝インフラ":"東芝",
        "ＪＦＥスチール": "JFE","ＪＦＥエンジニアリング": "JFE","ＪＦＥ建材":"JFE","ＪＦＥケミカル":"JFE","ＪＦＥテクノリサーチ":"JFE",
        "ＩＨＩプラント":"ＩＨＩ","ＩＨＩアグリテック":"ＩＨＩ","ＩＨＩエアロスペース":"ＩＨＩ","石川島芝浦機械":"ＩＨＩ",
        "旭化成イーマテリアルズ":"旭化成","旭化成ホームズ":"旭化成","旭化成メディカル":"旭化成","旭化成せんい":"旭化成","旭化成エレクトロニクス":"旭化成","旭化成ケミカルズ":"旭化成","セルガード":"旭化成",
        "三井Ｅ＆Ｓプラントエンジニアリング":"三井Ｅ＆Ｓ",
        "ゼネラルエレクトリックテクノロジーゲゼルシャフトミットベシュレンクテルハフツング":"ゼネラルエレクトリック",
        "エクソンモービルリサーチアンドエンジニアリング":"エクソンモービル","エクソンモービルアップストリームリサーチ":"エクソンモービル",
        "日揮グローバル":"日揮","日揮触媒化成":"日揮",
        "豊田中央研究所":"トヨタ","トヨタ自動車":"トヨタ","豊田自動織機":"トヨタ","トヨタモーターエンジニアリングアンドマニュファクチャリングノースアメリカインコーポレイティド":"トヨタ",
        "ヤンマー農機":"ヤンマー","ヤンマーパワーテクノロジー":"ヤンマー","ヤンマーエンジニアリング":"ヤンマー","ヤンマーアグリ":"ヤンマー","ヤンマーグリーンシステム":"ヤンマー","ヤンマーディーゼル":"ヤンマー","ヤンマーデイーゼル":"ヤンマー","ヤンマーノウキ":"ヤンマー",
        "神崎高級工機製作所":"ヤンマー","セイレイ工業":"ヤンマー",
        "ニコンビジネスサービス":"ニコン",
        "ユ—シン":"ユーシン","ユ－シン":"ユーシン",
        "リコ－":"リコー",
        "ロ－ベルトボツシユゲゼルシヤフトミツトベシユレンクテルハフツング":"ボッシュ","ボッシュオートモーティブシステム":"ボッシュ","ローベルトボツシユゲゼルシヤフトミツトベシユレンクテルハフツング":"ボッシュ","ローベルトボッシュゲゼルシャフトミットベシュレンクテルハフツング":"ボッシュ",
        "ＮＥＣエンベデッドプロダクツ":"ＮＥＣ","ＮＥＣシステムテクノロジー":"ＮＥＣ","ＮＥＣスペーステクノロジー":"ＮＥＣ","ＮＥＣソリューションイノベータ":"ＮＥＣ","ＮＥＣプラットフォームズ":"ＮＥＣ","ＮＥＣモバイルコミュニケーションズ":"ＮＥＣ","日本電気":"ＮＥＣ",
        "キヤノンマーケティングジャパン":"キヤノン","キヤノンオプトロン":"キヤノン",
        "クボタケミックス":"クボタ","クボタ機械サービス":"クボタ","クボタ鉄工":"クボタ","久保田鉄工":"クボタ",
        "シヤ－プ":"シャープ",
        "住友化学工業":"住友化学",
        "トプコンポジショニング":"トプコン",
        "ビーエーエスエフアグロトレードマークス":"BASF","ビーエーエスエフアグロベーブイ":"BASF","ビーエーエスエフソシエタスヨーロピア":"BASF","ベーアーエスエフアグロトレードマークス":"BASF","ベーアーエスエフエスエー":"BASF","ベーアーエスエフフューエルセル":"BASF","ビーエーエスエフ":"BASF",
        "ヒタチエナジーリミテッド":"日立",
        "マルコンデンシ":"マルコン電子",
        "日立エンジニアリング":"日立","日立ケーイー":"日立","日立ソリューションズ":"日立","日立ソリューションズ東日本":"日立","日立チャネルソリューションズ":"日立","日立ハイテクフィールディング":"日立","日立パワーソリューションズ":"日立","日立プラントテクノロジー":"日立","日立産機システム":"日立","日立照明":"日立","日立情報通信エンジニアリング":"日立","日立製作所":"日立","日立冷熱":"日立","日立東日本":"日立",
        "日立オートモティブ":"日立Ａｓｔｅｍｏ","日信工業":"日立Ａｓｔｅｍｏ",
        "日立金属":"プロテリアル","日立電線":"プロテリアル",
        "ソニーセミコンダクタソリューションズ":"ソニー","ソニーグループ":"ソニー",
        "富士重工業":"ＳＵＢＡＲＵ",
        "立石電機":"オムロン",
        "マルコン電子":"日本ケミコン",
        "佐竹製作所":"サタケ",
        "UBE":"宇部興産","ＵＢＥ":"宇部興産",
        "日本電池":"ユアサ","ジーエスユアサ":"ユアサ","ジーエスユアサインダストリー":"ユアサ","ＧＳユアサインダストリー":"ユアサ","ユアサ開発":"ＧＳユアサ","ＧＳユアサ":"ユアサ","ユアサインダストリー":"ユアサ","ユアサコーポレーション":"ユアサ","ユアサコーポレーシヨン":"ユアサ",
        "エルジーエナジーソリューション":"LG","エルジーケム":"LG","エルジーケムエルティーディ":"LG","エルジーケミカル":"LG","エルジーケーブル":"LG",
        "日本曹達":"トクヤマ","徳山曹達":"トクヤマ",
        "東洋曹達工業":"東ソー","東ソーエフテック":"東ソー","東ソーファインケム":"東ソー",
        "ＴＯＰＰＡＮ":"凸版印刷",
        "エンビジョンＡＥＳＣジャパン":"AESC","ＡＥＳＣジャパン":"AESC","エンビジョンＡＥＳＣエナジーデバイス":"AESC","ＮＥＣエナジーデバイス":"AESC",
        "三星電子":"サムスン","三星エスディアイ":"サムスン","三星電機":"サムスン","サムスンエレクトロニクス":"サムスン","サムスン日本研究所":"サムスン","サムスンエスディアイ":"サムスン",
        "レゾナックパッケージング":"レゾナック","日立化成":"レゾナック","昭和電工パッケージング":"レゾナック","昭和電工マテリアルズ":"レゾナック","昭和電工":"レゾナック","日立粉末冶金":"レゾナック","日立化成工業":"レゾナック","新神戸電機":"レゾナック",
        "ルノーエスアエス":"ルノー",
        "日立マクセル":"マクセル","日立マクセルエナジー":"マクセル","日立ハイテク":"マクセル","日立ハイテクファイン":"マクセル","日立ハイテクサイエンス":"マクセル","日立ハイテクノロジーズ":"マクセル","マクセルエナジー":"マクセル",
        "日立ビークルエナジー":"ビークルエナジージャパン",
        "ＦＤＫ鳥取":"ＦＤＫ","ＦＤＫトワイセル":"ＦＤＫ",
        "寧徳時代新能源科技股分":"ＣＡＴＬ","寧徳新能源科技":"ＣＡＴＬ","寧徳新能源科技":"ＣＡＴＬ","香港時代新能源科技":"ＣＡＴＬ",
        "日新電機":"住友電気工業","富山住友電工":"住友電気工業","住友電工ファインポリマー":"住友電気工業","住友精化":"住友化学","住友住友ベークライト":"住友化学",
        "東レバッテリーセパレータフィルム":"東レ","東レフィルム加工":"東レ","トーレアドバンストマテリアルズコリア":"東レ","東レ韓国":"東レ","東レ東燃機能膜合同会社":"東レ","東レ東燃機能膜":"東レ","東レファインケミカル":"東レ","東レコーテックス":"東レ",
        "比亜迪股分":"ＢＹＤ","ビーワイディー":"ＢＹＤ",
        "サイオンパワーコーポレイション":"シオンパワー","サイオンパワー":"シオンパワー",
        "リチウムエナジーアンドパワーゲゼルシャフトミットベシュレンクテルハフッングウントコンパニーコマンディトゲゼルシャフト":"リチウムエナジーアンドパワー","リチウムエナジージャパン":"リチウムエナジーアンドパワー","リチウムエナジーアンドパワーウントコーカーゲー":"リチウムエナジーアンドパワー",
        "旭硝子":"ＡＧＣ","ＡＧＣセイミケミカル":"ＡＧＣ",
        "東洋製罐グループ":"東洋製罐",
        "ポスコケミカル":"ポスコ","ポスコインコーポレーティッド":"ポスコ",
        "ＴｅｒａＷａｔｔＴｅｃｈｎｏｌｏｇｙ":"テラワットテクノロジー",
        "華為技術":"ファーウェイ",
        "テーデーカーエレクトロニクスアーゲー":"ＴＤＫ",
        "東洋紡フイルムソリューション":"東洋紡","東洋紡績":"東洋紡",
        "クアンタムスケイプバテリー":"クアンタム","クアンタムスケープバッテリー":"クアンタム",
        "イーアイデュポンドウヌムールアンド":"デュポン","デュポンセイフティーアンドコンストラクション":"デュポン","デュポンセーフティーアンドコンストラクション":"デュポン",
        "パナソニックＥＶエナジー": "プライムアースＥＶエナジー",
        "ソルベイスペシャルティポリマーズイタリーエスピーエー":"ソルベイ","ソルベイスペシャルティポリマーズユーエスエー":"ソルベイ","ソルベイアセトウ":"ソルベイ","ソルヴェイ（ソシエテアノニム）":"ソルベイ","ソルヴェイ":"ソルベイ",
        "ＪＸ金属サーキュラーソリューションズ":"ＥＮＥＯＳ","ＪＸＴＧエネルギー":"ＥＮＥＯＳ","ＪＸ日鉱日石エネルギー":"ＥＮＥＯＳ","ＥＮＥＯＳマテリアル":"ＥＮＥＯＳ","東燃ゼネラル石油":"ＥＮＥＯＳ","ＪＸエネルギー":"ＥＮＥＯＳ","東燃化学":"ＥＮＥＯＳ","ＪＸ金属":"ＥＮＥＯＳ","新日本石油":"ＥＮＥＯＳ","ＪＸ日鉱日石金属":"ＥＮＥＯＳ",
        "帝人化成":"帝人","帝人フロンティア":"帝人","帝人フィルム":"帝人","帝人ファイバー":"帝人","帝人テクノプロダクツ":"帝人","帝人株式会社":"帝人","帝人コードレ":"帝人","帝人知的財産センター":"帝人",
        "富士ゼロックス":"富士フイルム","富士フイルムビジネスイノベーション":"富士フイルム","富士フイルム和光純薬":"富士フイルム",
        "ＪＮＣ石油化学":"ＪＮＣ",
        "スリーエム":"３Ｍ","スリーエムイノベイティブプロパティズ":"３Ｍ","３Ｍイノベイティブプロパティズ":"３Ｍ",
        "シェンズェンカプチェムテクノロジー":"新宙邦科技","深せん新宙邦科技股ふん":"新宙邦科技",
        "ハイドローケベック":"ハイドロケベック","イドローケベック":"ハイドロケベック","ハイドローケベツク":"ハイドロケベック",
        "アーケマ":"アルケマ","アルケマフランス":"アルケマ","アルケマフランス":"アルケマ",
        "エイ１２３ライアビリティ":"万向集団","エー１２３":"万向集団","エー１２３ライアビリティ":"万向集団",
        "ソシエテデプロデユイネツスルソシエテアノニム":"ネスレ","ソシエテデプロデュイネスレエスアー":"ネスレ","ソシエテデプロデユイネツスレソシエテアノニム":"ネスレ",
        "ユニリーバーアイピーベーフェー":"ユニリーバ","ユニリーバーナームローゼベンノートシヤープ":"ユニリーバ","ユニリーバーナームローゼベンノートシャープ":"ユニリーバ",
        "ザプロクターエンドギャンブル":"Ｐ＆Ｇ","ザプロクターエンドギヤンブルカンパニー":"Ｐ＆Ｇ",
        "ジェイエスアール":"ＪＳＲ",
        "バイエルクロップサイエンス":"バイエル","バイエルインテレクチュアルプロパティゲゼルシャフトミットベシュレンクテルハフツング":"バイエル","バイエルクロップサイエンスエルピー":"バイエル","バイエルクロップサイエンスアクチェンゲゼルシャフト":"バイエル","バイエルビジネスサービシズゲゼルシャフトミットベシュレンクテルハフツング":"バイエル",
        "バイエルアクチエンゲゼルシャフト":"バイエル",
        "東芝ライテック":"東芝",
        "騰訊科技（深セン）":"テンセント","テンセントテクノロジー（シェンジェン）":"テンセント","シェンジェンテンセントコンピューター":"テンセント","テンセントアメリカ":"テンセント","深セン市騰訊計算机系統":"テンセント",
        "インターナショナルビジネスマシーンズ":"IBM",
        "グーグル":"Google",
        "日本電信電話":"NTT","エヌティーティーリサーチ":"NTT",
        "ベイジンバイドゥネットコムサイエンステクノロジー":"バイドゥ",
        "マイクロソフトテクノロジーライセンシング":"Microsoft",
        "アイキューエムフィンランドオイ":"IQM quantum","アイキューエムフィンランドオーワイ":"IQM quantum",
        "アリババグループ":"アリババ","アリババイノベーションプライベート":"アリババ","アリババ（チャイナ）":"アリババ",
        "イオンキュークアンタムカナダ":"IonQ","イオンキュー":"IonQ",
        "ワンキュービーインフォメーションテクノロジーズ":"1QBit","１キュービーインフォメーションテクノロジーズ":"1QBit",
        "ディーーウェイブ":"D-Wave",
        "イェールユニバーシティー":"イェール大学",
        "ユニバーシティオブメリーランドカレッジパーク":"メリーランド大学",
        "プレジデントアンドフェローズオブハーバードカレッジ":"ハーバード大学",
        "オックスフォードユニバーシティイノベーションリミティド":"オックスフォード大学","オックスフォードユニヴァーシティイノヴェーション":"オックスフォード大学","オックスフォードユニバーシティイノベーション":"オックスフォード大学",
        "科学技術振興機構":"JST",
        }

    # 該当列のデータを走査して、辞書に基づいて変換
    for row in range(2, sheet.max_row + 1):  # ヘッダーを飛ばして2行目から始める
        cell = sheet.cell(row=row, column=target_column)
        if cell.value:
            for old_name, new_name in replace_dict.items():
                if old_name in cell.value: # type: ignore
                    cell.value = cell.value.replace(old_name, new_name) # type: ignore

# ヘッダーとして探す値
header_value = '出願日'
target_col = None

# 最上部のセルをチェックして、目的の列を見つける
for col in sheet.iter_cols(min_row=1, max_row=1):
    if col[0].value == header_value:
        target_col = col[0].column
        break

if target_col is not None:
    sheet.insert_cols(target_col)  # 出願日の左側に列を挿入
    sheet.cell(row=1, column=target_col, value='出願年')  # 新しい列にヘッダーを設定

    # 出願日のデータから年を取り出し、挿入した列にコピーする
    for row in sheet.iter_rows(min_row=2, min_col=target_col + 1, max_col=target_col + 1):
        date_cell = row[0]  # 出願日のセルを取得
        if isinstance(date_cell.value, (int, float)):  # 数値として扱われている場合
            # Excelの日付シリアル値をdatetimeに変換
            serial_date = date_cell.value
            base_date = datetime(1899, 12, 30)  # Excelのシリアル値の基準日
            date_value = base_date + timedelta(days=serial_date)
        elif isinstance(date_cell.value, datetime):
            # 日付がすでにdatetime型の場合
            date_value = date_cell.value
        else:
            # 日付でない場合は空白にする
            date_value = None

        if date_value:
            year = date_value.year
            # 年を挿入した列に設定
            sheet.cell(row=date_cell.row, column=target_col, value=year) # type: ignore
        else:
            # 日付でない場合は空白にする
            sheet.cell(row=date_cell.row, column=target_col, value='')
  

header_value = '更新出願人・権利者氏名'
target_col = None

# 最上部のセルをチェックして、目的の列を見つける
for col in sheet.iter_cols(min_row=1, max_row=1):
    if col[0].value == header_value:
        target_col = col[0].column
        break

           
           
if target_col is not None:
    sheet.insert_cols(target_col)  # '更新出願人・権利者氏名'の左側に列を挿入
    sheet.cell(row=1, column=target_col, value='筆頭出願人')  # 新しい列にヘッダーを設定

    # '更新出願人・権利者氏名'のデータを分割し、最初の部分を新しい列にコピーする
    for row in sheet.iter_rows(min_row=2, min_col=target_col + 1, max_col=target_col + 1):
        original_value = row[0].value  # '更新出願人・権利者氏名'のセルを取得
        if original_value is not None and isinstance(original_value, str):
            # セルの値を ',' または '，' で分割
            split_values = original_value.split(',') if ',' in original_value else original_value.split('，')
            # 最初の部分を取得し、余分な文字を削除
            if split_values:
                head_applicant = split_values[0]
                # 不要な文字を削除
                head_applicant = head_applicant.replace(' ', '').replace('　', '').replace('・', '').replace('.', '')
            else:
                head_applicant = original_value
        else:
            head_applicant = original_value
        
        # 新しい列に設定
        sheet.cell(row=row[0].row, column=target_col, value=head_applicant) # type: ignore


# 公報IPCの列を探す
ipc_column_header = '公報IPC'
ipc_col = None
for col in sheet.iter_cols(1, sheet.max_column):
    if col[0].value == ipc_column_header:
        ipc_col = col[0].column
        break

if ipc_col is not None:
    # 公報IPCの左側に3つの列を挿入してヘッダーを設定
    sheet.insert_cols(ipc_col)
    sheet.cell(row=1, column=ipc_col, value='筆頭IPCメイングループ')

    sheet.insert_cols(ipc_col)
    sheet.cell(row=1, column=ipc_col, value='筆頭IPCサブクラス')

    sheet.insert_cols(ipc_col)
    sheet.cell(row=1, column=ipc_col, value='筆頭IPCサブグループ')

    # 公報IPCのデータから各部分を抽出して新しい列にコピーする
    for row in sheet.iter_rows(min_row=2, min_col=ipc_col + 3, max_col=ipc_col + 3):
        ipc_value = row[0].value  # 公報IPCのセルを取得
        if ipc_value is not None and isinstance(ipc_value, str):
            # 筆頭IPCメイングループ: /までの部分を取得
            main_group = ipc_value.split('/')[0] if '/' in ipc_value else ipc_value
            # 筆頭IPCサブクラス: 先頭から4文字を取得
            sub_class = ipc_value[:4]
            # 筆頭IPCサブグループ: , または ＠ の前までの部分を取得
            if ',' in ipc_value:
                sub_group = ipc_value.split(',')[0]
            elif '＠' in ipc_value:
                sub_group = ipc_value.split('＠')[0]
            else:
                sub_group = ipc_value
        else:
            main_group = sub_class = sub_group = ipc_value

        # 新しい列に設定
        sheet.cell(row=row[0].row, column=ipc_col, value=sub_group)  # type: ignore # 筆頭IPCサブグループ
        sheet.cell(row=row[0].row, column=ipc_col + 1, value=sub_class)  # type: ignore # 筆頭IPCサブクラス
        sheet.cell(row=row[0].row, column=ipc_col + 2, value=main_group)  # type: ignore # 筆頭IPCメイングループ

        
# 公報FIの列を探す
fi_column_header = '公報FI'
fi_col = None

for col in sheet.iter_cols(1, sheet.max_column):
    if col[0].value == fi_column_header:
        fi_col = col[0].column
        break

if fi_col is not None:
    # 公報FIの左側に3つの列を挿入してヘッダーを設定
    sheet.insert_cols(fi_col)  # 筆頭FIサブグループ
    sheet.cell(row=1, column=fi_col, value='筆頭FIメイングループ')

    sheet.insert_cols(fi_col)  # 筆頭FIサブクラス
    sheet.cell(row=1, column=fi_col, value='筆頭FIサブクラス')

    sheet.insert_cols(fi_col)  # 筆頭FIメイングループ
    sheet.cell(row=1, column=fi_col, value='筆頭FIサブグループ')

    # 公報FIのデータから各部分を抽出して新しい列にコピーする
    for row in sheet.iter_rows(min_row=2, min_col=fi_col + 4, max_col=fi_col + 4):
        fi_value = row[0].value  # 公報FIのセルを取得
        if fi_value is not None and isinstance(fi_value, str):
            # 筆頭FIメイングループ: /までの部分を取得
            main_group = fi_value.split('/')[0] if '/' in fi_value else fi_value
            # 筆頭FIサブクラス: 先頭から4文字を取得
            sub_class = fi_value[:4]
            # 筆頭FIサブグループ: , または ＠ の前までの部分を取得
            if ',' in fi_value:
                sub_group = fi_value.split(',')[0]
            elif '＠' in fi_value:
                sub_group = fi_value.split('＠')[0]
            else:
                sub_group = fi_value
        else:
            main_group = sub_class = sub_group = fi_value

        # 新しい列に設定
        sheet.cell(row=row[0].row, column=fi_col, value=main_group)  # type: ignore # 筆頭FIメイングループ
        sheet.cell(row=row[0].row, column=fi_col + 1, value=sub_class)  # type: ignore # 筆頭FIサブクラス
        sheet.cell(row=row[0].row, column=fi_col + 2, value=sub_group)  # type: ignore # 筆頭FIサブグループ
        
# 生死情報列の名前を見つける
column_life_death_header = '生死情報'
column_life_death = None


for col in sheet.iter_cols(1, sheet.max_column):
    if col[0].value == column_life_death_header:
        column_life_death = col[0].column
        break
    
if column_life_death is None:
    print("「生死情報」列が見つかりませんでした。")
else:
    # 「生死情報更新」を挿入（生死情報の左に1回だけ挿入）
    sheet.insert_cols(column_life_death)
    sheet.cell(row=1, column=column_life_death).value = '生死情報更新'

    # 生死情報の更新を挿入
    for row in range(2, sheet.max_row + 1):  # ヘッダー行をスキップ
        life_death_value = sheet.cell(row=row, column=column_life_death + 1).value  # 元の「生死情報」列は1つ右にずれる
        if life_death_value:
            # 「公開:」「死:」「登録:」などのケースに対応
            parts = life_death_value.split(':', 1) # type: ignore
            if len(parts) == 2 and parts[0] == parts[1]:
                new_value = parts[0]  # 前後が同じ場合は片方を残す
            elif life_death_value.startswith('公開:'): # type: ignore
                new_value = parts[1]  # 公開:の後を残す
            elif life_death_value.startswith('死:'): # type: ignore
                new_value = '死'  # 死:の場合は死のみ
            elif life_death_value.startswith('登録:'): # type: ignore
                new_value = '登録'  # 登録:の場合は登録のみ
            else:
                new_value = life_death_value  # それ以外はそのまま
        else:
            new_value = None  # Noneや空白の値に対応

        # 「生死情報更新」列に新しい値を挿入
        sheet.cell(row=row, column=column_life_death).value = new_value
wb.save(fname)

# ワークブックを閉じる
wb.close()

