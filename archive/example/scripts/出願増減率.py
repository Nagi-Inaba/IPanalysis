import openpyxl
from collections import defaultdict
import sys 


# ワークブックをロード
fname = sys.argv[1]
target_year = int(sys.argv[2])

# 前後10年の範囲
year_range = 10

# Excelファイルを開く
wb = openpyxl.load_workbook(fname)
sheet = wb['データ']

# 列番号を取得
applicant_column = None
year_column = None

# 列名から列番号を取得
for col in sheet.iter_cols(1, sheet.max_column):
    if col[0].value == '更新出願人・権利者氏名':
        applicant_column = col[0].column
    elif col[0].value == '出願年':
        year_column = col[0].column
    
    if applicant_column and year_column:
        break

if applicant_column is None or year_column is None:
    print("必要な列が見つかりません。")
    exit()

# データを集計
data = defaultdict(lambda: defaultdict(int))
total_data = defaultdict(int)

for row in sheet.iter_rows(min_row=2):
    # 行データの取得
    row_data = [cell.value for cell in row]

    # データが不足している場合のチェック
    if len(row_data) < max(applicant_column, year_column):
        print("行データが不完全です。")
        continue

    year_value = row_data[year_column - 1]
    applicant_value = row_data[applicant_column - 1]

    # 年と出願人情報が空白でないことを確認
    if year_value is not None and applicant_value is not None:
        # 年が整数か確認
        if isinstance(year_value, int):
            # 出願人名が複数ある場合、カンマで分割し、同じセル内で重複しないようにセットを使用
            applicants = set([app.strip() for app in applicant_value.split(',')]) # type: ignore

            for applicant in applicants:
                total_data[applicant] += 1  # 総数をカウント

                # 基準年の前後10年のデータを分類
                if target_year - year_range <= year_value < target_year:
                    data['before_10'][applicant] += 1  # 基準年より前の10年
                elif target_year <= year_value < target_year + 5:
                    data['after_5_first'][applicant] += 1  # 基準年から5年以内
                elif target_year + 5 <= year_value <= target_year + year_range:
                    data['after_5_second'][applicant] += 1  # 基準年から5年以上10年以内

# 結果を計算
results = []
for applicant in total_data.keys():
    before_10_count = data['before_10'].get(applicant, 0)
    after_5_first_count = data['after_5_first'].get(applicant, 0)
    after_5_second_count = data['after_5_second'].get(applicant, 0)
    after_10_total = after_5_first_count + after_5_second_count
    total_20_year_count = before_10_count + after_10_total  # 前後20年の合計

    # 前10年 vs 後10年の増減率 (分母は前後20年の合計)
    if total_20_year_count == 0:
        percentage_change_10 = 0
    else:
        percentage_change_10 = (after_10_total - before_10_count) / total_20_year_count 

    # 最初の5年 vs 次の5年 (分母は後半の10年の合計)
    if (after_5_first_count + after_5_second_count) == 0:
        percentage_change_second_5 = 0
    else:
        percentage_change_second_5 = (after_5_second_count - after_5_first_count) / (after_5_first_count + after_5_second_count) 

    results.append({
        'Applicant': applicant,
        'Before 10 years': before_10_count,
        'After 10 years': after_10_total,
        'After 5 years (first)': after_5_first_count,
        'After 5 years (second)': after_5_second_count,
        'Total': total_20_year_count,
        'Percentage Change (Before 10 vs After 10)': percentage_change_10,
        'Percentage Change (First 5 vs Second 5)': percentage_change_second_5
    })

# 新しいシートを作成し、結果を記述
result_sheet = wb.create_sheet(title=f"出願人増減率_{target_year}")

# ヘッダーの設定（基準年に応じてヘッダーをカスタマイズ）
headers = [
    '出願人', 
    f'{target_year - year_range}～{target_year}以前',  # 基準年より前10年
    f'{target_year}～{target_year + year_range}以降',  # 基準年から後10年
    f'{target_year}～{target_year + 5}前5年',  # 基準年から5年以内
    f'{target_year + 5}～{target_year + year_range}後5年',  # 基準年から5年以上10年以内
    '増減率 (前10年 vs 後10年)', 
    '増減率 (最初の5年 vs 次の5年)',
    '出願件数',
]

for col, header in enumerate(headers, start=1):
    result_sheet.cell(row=1, column=col, value=header)
# データの書き込み
for row_index, result in enumerate(results, start=2):
    result_sheet.cell(row=row_index, column=1, value=result['Applicant'])
    result_sheet.cell(row=row_index, column=2, value=result['Before 10 years'])
    result_sheet.cell(row=row_index, column=3, value=result['After 10 years'])
    result_sheet.cell(row=row_index, column=4, value=result['After 5 years (first)'])
    result_sheet.cell(row=row_index, column=5, value=result['After 5 years (second)'])
    result_sheet.cell(row=row_index, column=6, value=result['Percentage Change (Before 10 vs After 10)'])
    result_sheet.cell(row=row_index, column=7, value=result['Percentage Change (First 5 vs Second 5)'])
    result_sheet.cell(row=row_index, column=8, value=result['Total'])

# 結果を保存
wb.save(fname)
print(f"増減率の結果が新しいシート '出願人増減率_{target_year}' に書き込まれました。")
