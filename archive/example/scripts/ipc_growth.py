import openpyxl
from collections import defaultdict
import sys 


# ワークブックをロード
fname = sys.argv[1]
target_year = int(sys.argv[2])

# 前後10年の範囲
year_range = 10

# Excelファイルを開く
wb = openpyxl.load_workbook(fname)
sheet = wb['データ']

# 列番号を取得
ipc_column = None
year_column = None

# 列名から列番号を取得
for col in sheet.iter_cols(1, sheet.max_column):
    if col[0].value == '公報IPC':
        ipc_column = col[0].column
    elif col[0].value == '出願年':
        year_column = col[0].column
    
    if ipc_column and year_column:
        break

if ipc_column is None or year_column is None:
    print("必要な列が見つかりません。")
    exit()

# データを集計
data = defaultdict(lambda: defaultdict(int))
total_data = defaultdict(int)

for row in sheet.iter_rows(min_row=2):
    # 行データの取得と確認
    row_data = [cell.value for cell in row]

    # データが不足している場合
    if len(row_data) < max(ipc_column, year_column):
        print("行データが不完全です。")
        continue

    year_value = row_data[year_column - 1]
    ipc_value = row_data[ipc_column - 1]

    # データが空白でないことを確認
    if year_value is not None and ipc_value is not None:
        # IPCが複数の場合、区切り文字で分割
        ipc_list = ipc_value.split(',') if isinstance(ipc_value, str) else [ipc_value]

        # 年が整数であることを確認
        if isinstance(year_value, int):
            for ipc in ipc_list:
                ipc = ipc.strip()  # type: ignore # 余分な空白を除去
                total_data[ipc] += 1

                # 前後10年の集計
                if target_year - year_range <= year_value < target_year:
                    data['before_10'][ipc] += 1  # 基準年より前の10年
                elif target_year <= year_value < target_year + 5:
                    data['after_5_first'][ipc] += 1  # 基準年から5年以内
                elif target_year + 5 <= year_value <= target_year + year_range:
                    data['after_5_second'][ipc] += 1  # 基準年から5年以上10年以内

# 結果を計算
results = []
for ipc in total_data.keys():
    before_10_count = data['before_10'].get(ipc, 0)  # 前10年のデータ
    after_5_first_count = data['after_5_first'].get(ipc, 0)  # 後半10年の最初の5年のデータ
    after_5_second_count = data['after_5_second'].get(ipc, 0)  # 後半10年の次の5年のデータ
    after_10_total = after_5_first_count + after_5_second_count  # 後半10年の合計
    total_20_year_count = before_10_count + after_10_total  # 前10年と後10年の合計

    # 前10年 vs 後10年の増減率
    if total_20_year_count == 0:
        percentage_change_10 = 0
    else:
        # 分母は前10年と後10年の合計
        percentage_change_10 = (after_10_total - before_10_count) / total_20_year_count 

    # 最初の5年 vs 次の5年の増減率
    if (after_5_first_count + after_5_second_count) == 0:
        percentage_change_second_5 = 0
    else:
        # 分母は後半10年の合計
        percentage_change_second_5 = (after_5_second_count - after_5_first_count) / (after_5_first_count + after_5_second_count) 

    results.append({
        'IPC': ipc,
        'Before 10 years': before_10_count,
        'After 10 years': after_10_total,
        'After 5 years (first)': after_5_first_count,
        'After 5 years (second)': after_5_second_count,
        'Total': total_20_year_count,
        'Percentage Change (Before 10 vs After 10)': percentage_change_10,
        'Percentage Change (First 5 vs Second 5)': percentage_change_second_5
    })

# 新しいシートを作成し、結果を記述
result_sheet = wb.create_sheet(title=f"公報IPC増減率_{target_year}")

# ヘッダーの設定（基準年に応じてヘッダーをカスタマイズ）
headers = [
    'IPC', 
    f'{target_year - year_range}～{target_year}',  # 基準年より前10年
    f'{target_year}～{target_year + year_range}',  # 基準年から後10年
    f'{target_year}～{target_year + 5}前5年',  # 基準年から5年以内
    f'{target_year + 5}～{target_year + year_range}後5年',  # 基準年から5年以上10年以内
    '10年増減率 ', 
    '5年増減率 ',
    '出願件数', 
]

for col, header in enumerate(headers, start=1):
    result_sheet.cell(row=1, column=col, value=header)

# データの書き込み
for row_index, result in enumerate(results, start=2):
    result_sheet.cell(row=row_index, column=1, value=result['IPC'])
    result_sheet.cell(row=row_index, column=2, value=result['Before 10 years'])
    result_sheet.cell(row=row_index, column=3, value=result['After 10 years'])
    result_sheet.cell(row=row_index, column=4, value=result['After 5 years (first)'])
    result_sheet.cell(row=row_index, column=5, value=result['After 5 years (second)'])
    result_sheet.cell(row=row_index, column=8, value=result['Total'])
    result_sheet.cell(row=row_index, column=6, value=result['Percentage Change (Before 10 vs After 10)'])
    result_sheet.cell(row=row_index, column=7, value=result['Percentage Change (First 5 vs Second 5)'])

# 結果を保存
wb.save(fname)
print(f"増減率の結果が新しいシート '公報IPC増減率_{target_year}' に書き込まれました。")
