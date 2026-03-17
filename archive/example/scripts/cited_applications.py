import openpyxl
import re
import sys 


# ワークブックをロード
fname = sys.argv[1]

# Excelファイルを開く
wb = openpyxl.load_workbook(fname)
sheet = wb['データ']  # データシートを指定してください

# 列番号を取得
applicant_column = None
application_number_column = None
application_year_column = None
citation_count_column = None
life_death_column = None  # 生死情報更新の列

# 列名から列番号を取得
for col in sheet.iter_cols(1, sheet.max_column):
    if col[0].value == '筆頭出願人':
        applicant_column = col[0].column
    elif col[0].value == '出願番号':
        application_number_column = col[0].column
    elif col[0].value == '出願年':
        application_year_column = col[0].column
    elif col[0].value == '被引用回数':
        citation_count_column = col[0].column
    elif col[0].value == '生死情報更新':
        life_death_column = col[0].column
    
    if all([applicant_column, application_number_column, application_year_column, citation_count_column, life_death_column]):
        break

if not all([applicant_column, application_number_column, application_year_column, citation_count_column, life_death_column]):
    print("必要な列が見つかりません。")
    exit()

# データを集計
citation_data = []

for row in sheet.iter_rows(min_row=2):
    applicant = row[applicant_column - 1].value  # type: ignore # 筆頭出願人
    application_number = row[application_number_column - 1].value  # type: ignore # 出願番号
    application_year = row[application_year_column - 1].value  # type: ignore # 出願年
    citation_str = row[citation_count_column - 1].value  # type: ignore # 被引用回数
    life_death_info = row[life_death_column - 1].value  # type: ignore # 生死情報更新

    # 被引用回数の処理
    if citation_str is not None:
        match = re.search(r'引用：(\d+)', citation_str)  # type: ignore # 引用の数字を抽出
        citation_count = int(match.group(1)) if match else 0
    else:
        citation_count = 0

    # 出願人、出願番号、出願年が存在する場合のみ処理
    if applicant and application_number and application_year:
        citation_data.append({
            'applicant': applicant,
            'application_number': application_number,
            'application_year': application_year,
            'citation_count': citation_count,
            'life_death_info': life_death_info  # 生死情報更新を追加
        })

# 被引用回数の多い順に並べ替え
citation_data.sort(key=lambda x: x['citation_count'], reverse=True)

# 新しいシートを作成し、結果を記述
result_sheet = wb.create_sheet(title="引用出願")

# ヘッダーの設定
headers = ['筆頭出願人 - 出願番号 - 生死情報更新', '出願番号', '出願年', '被引用回数']
for col, header in enumerate(headers, start=1):
    result_sheet.cell(row=1, column=col, value=header)

# データを書き込み
for row_index, data in enumerate(citation_data, start=2):
    # 出願人、出願番号、生死情報更新を同じセルに結合
    combined_value = f"{data['applicant']} - {data['application_number']}"
    if data['life_death_info']:  # 生死情報が存在する場合は追加
        combined_value += f" - {data['life_death_info']}"
    
    result_sheet.cell(row=row_index, column=1, value=combined_value)  # 出願人と出願番号と生死情報を同じセルに入れる
    result_sheet.cell(row=row_index, column=2, value=data['application_number'])  # 出願番号
    result_sheet.cell(row=row_index, column=3, value=data['application_year'])  # 出願年
    result_sheet.cell(row=row_index, column=4, value=data['citation_count'])  # 被引用回数

# 結果を保存
wb.save(fname)
print("被引用回数の多い出願が新しいシートに書き込まれました。")
