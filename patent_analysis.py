from __future__ import annotations

import argparse
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
import re
from typing import Any, Dict, List, Optional, Sequence, Tuple

import openpyxl


class AnalysisError(ValueError):
    """Raised when workbook content is invalid for analysis."""


@dataclass(frozen=True)
class AnalysisConfig:
    target_year: int
    year_range: int = 10
    source_sheet_name: str = "データ"
    ipc_column_name: str = "公報IPC"
    year_column_name: str = "出願年"
    result_sheet_prefix: str = "公報IPC増減率"


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("　", " ").strip()


def _find_required_column_indexes(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    config: AnalysisConfig,
) -> Tuple[int, int]:
    header_map: Dict[str, int] = {}
    for column_index, cell in enumerate(sheet[1], start=1):
        header = _normalize_header(cell.value)
        if header:
            header_map[header] = column_index

    ipc_index = header_map.get(config.ipc_column_name)
    year_index = header_map.get(config.year_column_name)
    if ipc_index is None or year_index is None:
        raise AnalysisError(
            f"必要な列が見つかりません。列名 '{config.ipc_column_name}' と "
            f"'{config.year_column_name}' を確認してください。"
        )
    return ipc_index, year_index


def _split_ipc_codes(raw_value: Any) -> List[str]:
    if raw_value is None:
        return []

    if isinstance(raw_value, str):
        tokens = re.split(r"[,\u3001\uFF0C;\uFF1B/\n\r\t]+", raw_value)
        return [token.strip() for token in tokens if token and token.strip()]

    token = str(raw_value).strip()
    return [token] if token else []


def _to_year(raw_value: Any) -> Optional[int]:
    if raw_value is None or isinstance(raw_value, bool):
        return None

    if isinstance(raw_value, datetime):
        return raw_value.year
    if isinstance(raw_value, date):
        return raw_value.year
    if isinstance(raw_value, int):
        return raw_value
    if isinstance(raw_value, float) and raw_value.is_integer():
        return int(raw_value)

    if isinstance(raw_value, str):
        text = raw_value.strip()
        if not text:
            return None

        matched = re.search(r"(19|20)\d{2}", text)
        if matched:
            return int(matched.group(0))

        try:
            numeric = float(text)
        except ValueError:
            return None
        return int(numeric) if numeric.is_integer() else None

    return None


def _validate_config(config: AnalysisConfig) -> None:
    if config.target_year < 1800 or config.target_year > 2200:
        raise AnalysisError(f"target_year={config.target_year} は範囲外です。")
    if config.year_range < 5:
        raise AnalysisError("year_range は 5 以上を指定してください。")


def _build_unique_sheet_name(
    workbook: openpyxl.workbook.workbook.Workbook,
    base_name: str,
) -> str:
    max_length = 31
    candidate = base_name[:max_length]
    counter = 2

    while candidate in workbook.sheetnames:
        suffix = f"_{counter}"
        candidate = f"{base_name[: max_length - len(suffix)]}{suffix}"
        counter += 1

    return candidate


def _write_result_sheet(
    workbook: openpyxl.workbook.workbook.Workbook,
    rows: List[Dict[str, Any]],
    config: AnalysisConfig,
) -> str:
    result_name = _build_unique_sheet_name(
        workbook,
        f"{config.result_sheet_prefix}_{config.target_year}",
    )
    sheet = workbook.create_sheet(title=result_name)

    before_start = config.target_year - config.year_range
    before_end = config.target_year - 1
    after_end = config.target_year + config.year_range
    first_end = config.target_year + 4
    second_start = config.target_year + 5

    headers = [
        "IPC",
        f"{before_start}～{before_end}",
        f"{config.target_year}～{after_end}",
        f"{config.target_year}～{first_end} (前5年)",
        f"{second_start}～{after_end} (後5年)",
        "10年増減率",
        "5年増減率",
        "出願件数",
    ]
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=column_index, value=header)

    for row_index, row in enumerate(rows, start=2):
        sheet.cell(row=row_index, column=1, value=row["IPC"])
        sheet.cell(row=row_index, column=2, value=row["before_count"])
        sheet.cell(row=row_index, column=3, value=row["after_count"])
        sheet.cell(row=row_index, column=4, value=row["after_first_5_count"])
        sheet.cell(row=row_index, column=5, value=row["after_second_5_count"])
        cell_10 = sheet.cell(row=row_index, column=6, value=row["pct_change_10"])
        cell_5 = sheet.cell(row=row_index, column=7, value=row["pct_change_second_5"])
        sheet.cell(row=row_index, column=8, value=row["total_count"])
        cell_10.number_format = "0.00%"
        cell_5.number_format = "0.00%"

    sheet.freeze_panes = "A2"
    sheet.column_dimensions["A"].width = 22
    sheet.column_dimensions["B"].width = 16
    sheet.column_dimensions["C"].width = 16
    sheet.column_dimensions["D"].width = 18
    sheet.column_dimensions["E"].width = 18
    sheet.column_dimensions["F"].width = 12
    sheet.column_dimensions["G"].width = 12
    sheet.column_dimensions["H"].width = 12

    return result_name


def analyze_workbook(
    workbook: openpyxl.workbook.workbook.Workbook,
    config: AnalysisConfig,
) -> Tuple[List[Dict[str, Any]], str]:
    _validate_config(config)

    if config.source_sheet_name not in workbook.sheetnames:
        raise AnalysisError(
            f"シート '{config.source_sheet_name}' が見つかりません。"
        )

    source_sheet = workbook[config.source_sheet_name]
    ipc_index, year_index = _find_required_column_indexes(source_sheet, config)
    required_width = max(ipc_index, year_index)

    counts: Dict[str, Dict[str, int]] = defaultdict(
        lambda: {"before": 0, "after_first": 0, "after_second": 0}
    )
    totals: Dict[str, int] = defaultdict(int)

    for row in source_sheet.iter_rows(min_row=2, values_only=True):
        if row is None or len(row) < required_width:
            continue

        year_value = _to_year(row[year_index - 1])
        ipc_codes = _split_ipc_codes(row[ipc_index - 1])
        if year_value is None or not ipc_codes:
            continue

        for ipc in ipc_codes:
            totals[ipc] += 1
            ipc_counts = counts[ipc]

            if config.target_year - config.year_range <= year_value < config.target_year:
                ipc_counts["before"] += 1
            elif config.target_year <= year_value < config.target_year + 5:
                ipc_counts["after_first"] += 1
            elif config.target_year + 5 <= year_value <= config.target_year + config.year_range:
                ipc_counts["after_second"] += 1

    rows: List[Dict[str, Any]] = []
    for ipc in sorted(totals.keys(), key=lambda code: (-totals[code], code)):
        before_count = counts[ipc]["before"]
        after_first = counts[ipc]["after_first"]
        after_second = counts[ipc]["after_second"]
        after_total = after_first + after_second
        total_count = before_count + after_total

        pct_change_10 = (
            (after_total - before_count) / total_count if total_count else 0.0
        )
        pct_change_second_5 = (
            (after_second - after_first) / after_total if after_total else 0.0
        )

        rows.append(
            {
                "IPC": ipc,
                "before_count": before_count,
                "after_count": after_total,
                "after_first_5_count": after_first,
                "after_second_5_count": after_second,
                "pct_change_10": pct_change_10,
                "pct_change_second_5": pct_change_second_5,
                "total_count": total_count,
            }
        )

    result_sheet_name = _write_result_sheet(workbook, rows, config)
    return rows, result_sheet_name


def analyze_excel_bytes(
    file_bytes: bytes,
    config: AnalysisConfig,
) -> Tuple[List[Dict[str, Any]], bytes, str]:
    workbook = openpyxl.load_workbook(filename=BytesIO(file_bytes))
    rows, result_sheet_name = analyze_workbook(workbook, config)
    output = BytesIO()
    workbook.save(output)
    return rows, output.getvalue(), result_sheet_name


def analyze_excel_file(
    input_file: Path,
    config: AnalysisConfig,
    output_file: Optional[Path] = None,
) -> Tuple[List[Dict[str, Any]], str, Path]:
    if not input_file.exists():
        raise FileNotFoundError(f"入力ファイルが見つかりません: {input_file}")

    workbook = openpyxl.load_workbook(filename=str(input_file))
    rows, result_sheet_name = analyze_workbook(workbook, config)

    resolved_output = output_file
    if resolved_output is None:
        resolved_output = input_file.with_name(f"{input_file.stem}_analysis.xlsx")

    workbook.save(str(resolved_output))
    return rows, result_sheet_name, resolved_output


def build_cli_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="IPCの時系列増減率を計算してExcelに結果シートを追加します。"
    )
    parser.add_argument("input_file", help="入力Excelファイル (.xlsx)")
    parser.add_argument("target_year", type=int, help="基準年")
    parser.add_argument(
        "--output-file",
        type=Path,
        default=None,
        help="出力ファイルパス（省略時: *_analysis.xlsx）",
    )
    parser.add_argument("--year-range", type=int, default=10, help="比較レンジ（年）")
    parser.add_argument(
        "--source-sheet",
        default="データ",
        help="分析対象シート名",
    )
    parser.add_argument(
        "--ipc-column",
        default="公報IPC",
        help="IPC列のヘッダー名",
    )
    parser.add_argument(
        "--year-column",
        default="出願年",
        help="出願年列のヘッダー名",
    )
    parser.add_argument(
        "--sheet-prefix",
        default="公報IPC増減率",
        help="結果シート名の接頭辞",
    )
    return parser


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = build_cli_parser()
    args = parser.parse_args(argv)

    config = AnalysisConfig(
        target_year=args.target_year,
        year_range=args.year_range,
        source_sheet_name=args.source_sheet,
        ipc_column_name=args.ipc_column,
        year_column_name=args.year_column,
        result_sheet_prefix=args.sheet_prefix,
    )

    try:
        rows, sheet_name, output_path = analyze_excel_file(
            input_file=Path(args.input_file),
            config=config,
            output_file=args.output_file,
        )
    except (AnalysisError, FileNotFoundError) as exc:
        print(f"[ERROR] {exc}")
        return 1

    print(f"分析完了: {len(rows)}件のIPCを集計しました。")
    print(f"結果シート: {sheet_name}")
    print(f"出力ファイル: {output_path}")
    return 0

